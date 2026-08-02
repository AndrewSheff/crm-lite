"""Сервис экспорта/импорта — Excel для клиентов и сделок с лимитом записей."""

import io

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload

from app.core.logging import get_logger
from app.models.client import Client
from app.models.deal import Deal
from app.models.user import User
from app.schemas.export import ImportResult

log = get_logger(__name__)


EXPORT_LIMIT = 10000


async def export_clients(user: User, db: AsyncSession) -> io.BytesIO:
    """Экспорт клиентов в Excel — лимит 10000 записей."""
    query = select(Client).options(joinedload(Client.owner), joinedload(Client.tags))

    if user.role != "admin":
        query = query.where(Client.owner_id == user.id)

    query = query.order_by(Client.created_at.desc()).limit(EXPORT_LIMIT)
    result = await db.execute(query)
    clients = result.unique().scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Клиенты"

    # заголовки
    headers = [
        "Имя", "Email", "Телефон", "Компания", "Отрасль",
        "Статус", "Источник", "Менеджер", "Теги", "Дата создания",
    ]
    ws.append(headers)

    for client in clients:
        tags = ", ".join(t.name for t in client.tags) if client.tags else ""
        ws.append([
            client.name,
            client.email or "",
            client.phone or "",
            client.company or "",
            client.industry or "",
            client.status,
            client.source or "",
            client.owner.name if client.owner else "",
            tags,
            client.created_at.strftime("%Y-%m-%d %H:%M"),
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


async def import_clients(
    file_content: bytes, user: User, db: AsyncSession
) -> dict:
    """Импорт клиентов из Excel — парсим и создаем."""
    try:
        wb = load_workbook(io.BytesIO(file_content))
    except Exception:
        return ImportResult(errors=["Не удалось прочитать файл"]).model_dump()

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    created = 0
    skipped = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        if not row or not row[0]:
            skipped += 1
            continue

        name = str(row[0]).strip()
        if not name:
            skipped += 1
            continue

        try:
            client = Client(
                name=name,
                email=str(row[1]).strip() if len(row) > 1 and row[1] else None,
                phone=str(row[2]).strip() if len(row) > 2 and row[2] else None,
                company=str(row[3]).strip() if len(row) > 3 and row[3] else None,
                industry=str(row[4]).strip() if len(row) > 4 and row[4] else None,
                status="lead",
                source="Импорт",
                owner_id=user.id,
            )
            db.add(client)
            created += 1
        except (ValueError, IntegrityError) as e:
            errors.append(f"Строка {i}: {e}")

    if created > 0:
        await db.flush()

    log.info("clients_imported", created=created, skipped=skipped)
    return ImportResult(created=created, skipped=skipped, errors=errors).model_dump()


async def export_deals(user: User, db: AsyncSession) -> io.BytesIO:
    """Экспорт сделок в Excel."""
    query = select(Deal).options(
        joinedload(Deal.client),
        joinedload(Deal.stage),
        joinedload(Deal.owner),
    )

    if user.role != "admin":
        query = query.where(Deal.owner_id == user.id)

    query = query.order_by(Deal.created_at.desc()).limit(EXPORT_LIMIT)
    result = await db.execute(query)
    deals = result.unique().scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Сделки"

    headers = [
        "Название", "Клиент", "Стадия", "Сумма", "Валюта",
        "Приоритет", "Вероятность", "Менеджер", "Дата создания", "Дата закрытия",
    ]
    ws.append(headers)

    for deal in deals:
        ws.append([
            deal.title,
            deal.client.name if deal.client else "",
            deal.stage.name if deal.stage else "",
            float(deal.amount) if deal.amount else 0,
            deal.currency,
            deal.priority,
            deal.probability,
            deal.owner.name if deal.owner else "",
            deal.created_at.strftime("%Y-%m-%d %H:%M"),
            deal.closed_at.strftime("%Y-%m-%d %H:%M") if deal.closed_at else "",
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
